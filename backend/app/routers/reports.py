from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from app.dependencies import get_current_user
from app.supabase_client import get_supabase_client
from pydantic import BaseModel
from typing import Optional, List
from datetime import date
import io
import csv

router = APIRouter()


class ReportRequest(BaseModel):
    report_type: str  # 'environmental', 'social', 'governance', 'esg_summary', 'custom'
    format: str = "csv"  # 'csv', 'excel', 'pdf'
    department_id: Optional[str] = None
    date_from: Optional[str] = None
    date_to: Optional[str] = None
    module: Optional[str] = None
    employee_id: Optional[str] = None
    challenge_id: Optional[str] = None
    category: Optional[str] = None


@router.post("/generate")
async def generate_report(
    request: ReportRequest,
    current_user: dict = Depends(get_current_user),
):
    """Generate a report with various filters and export formats."""
    supabase = get_supabase_client()

    if request.report_type == "environmental":
        data = await _get_environmental_data(supabase, request)
        title = "Environmental Report"
    elif request.report_type == "social":
        data = await _get_social_data(supabase, request)
        title = "Social Report"
    elif request.report_type == "governance":
        data = await _get_governance_data(supabase, request)
        title = "Governance Report"
    elif request.report_type == "esg_summary":
        data = await _get_esg_summary(supabase, request)
        title = "ESG Summary Report"
    else:
        data = await _get_custom_data(supabase, request)
        title = "Custom ESG Report"

    if request.format == "csv":
        return _generate_csv(data, title)
    elif request.format == "excel":
        return _generate_excel(data, title)
    elif request.format == "pdf":
        return _generate_pdf(data, title)
    elif request.format == "json":
        return {"data": data, "title": title}
    else:
        raise HTTPException(status_code=400, detail="Unsupported format")


async def _get_environmental_data(supabase, request):
    query = supabase.table("carbon_transactions").select(
        "*, departments(name), emission_factors(source_type, unit)"
    )
    if request.department_id:
        query = query.eq("department_id", request.department_id)
    if request.date_from:
        query = query.gte("date", request.date_from)
    if request.date_to:
        query = query.lte("date", request.date_to)
    result = query.order("date", desc=True).execute()

    rows = []
    for r in result.data:
        rows.append({
            "Date": r.get("date", ""),
            "Department": r.get("departments", {}).get("name", "") if r.get("departments") else "",
            "Source Type": r.get("source_type", ""),
            "Quantity": r.get("quantity", 0),
            "Emission (kg CO2e)": r.get("calculated_emission", 0),
            "Auto Calculated": "Yes" if r.get("auto_calculated") else "No",
        })
    return rows


async def _get_social_data(supabase, request):
    query = supabase.table("employee_participations").select(
        "*, profiles(full_name), csr_activities(title, date)"
    )
    if request.employee_id:
        query = query.eq("employee_id", request.employee_id)
    result = query.order("created_at", desc=True).execute()

    rows = []
    for r in result.data:
        rows.append({
            "Employee": r.get("profiles", {}).get("full_name", "") if r.get("profiles") else "",
            "Activity": r.get("csr_activities", {}).get("title", "") if r.get("csr_activities") else "",
            "Status": r.get("approval_status", ""),
            "Points Earned": r.get("points_earned", 0),
            "Completion Date": r.get("completion_date", ""),
        })
    return rows


async def _get_governance_data(supabase, request):
    query = supabase.table("compliance_issues").select("*, audits(title), profiles(full_name)")
    if request.department_id:
        audit_query = supabase.table("audits").select("id").eq("department_id", request.department_id).execute()
        audit_ids = [a["id"] for a in audit_query.data]
        if audit_ids:
            query = query.in_("audit_id", audit_ids)
    result = query.order("due_date").execute()

    rows = []
    for r in result.data:
        rows.append({
            "Audit": r.get("audits", {}).get("title", "") if r.get("audits") else "",
            "Severity": r.get("severity", ""),
            "Description": r.get("description", ""),
            "Owner": r.get("profiles", {}).get("full_name", "") if r.get("profiles") else "",
            "Due Date": r.get("due_date", ""),
            "Status": r.get("status", ""),
        })
    return rows


async def _get_esg_summary(supabase, request):
    scores = supabase.table("department_scores").select("*, departments(name)").execute()
    rows = []
    for r in scores.data:
        rows.append({
            "Department": r.get("departments", {}).get("name", "") if r.get("departments") else "",
            "Environmental Score": r.get("environmental_score", 0),
            "Social Score": r.get("social_score", 0),
            "Governance Score": r.get("governance_score", 0),
            "Total Score": r.get("total_score", 0),
        })
    return rows


async def _get_custom_data(supabase, request):
    """Combine data based on selected module/filters."""
    all_data = []
    if not request.module or request.module == "environmental":
        all_data.extend(await _get_environmental_data(supabase, request))
    if not request.module or request.module == "social":
        all_data.extend(await _get_social_data(supabase, request))
    if not request.module or request.module == "governance":
        all_data.extend(await _get_governance_data(supabase, request))
    return all_data


def _generate_csv(data: list, title: str):
    if not data:
        return StreamingResponse(
            io.BytesIO(b"No data found"),
            media_type="text/csv",
            headers={"Content-Disposition": f'attachment; filename="{title}.csv"'}
        )

    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=data[0].keys())
    writer.writeheader()
    writer.writerows(data)

    content = output.getvalue().encode("utf-8")
    return StreamingResponse(
        io.BytesIO(content),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{title}.csv"'}
    )


def _generate_excel(data: list, title: str):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]

    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1B5E20", end_color="1B5E20", fill_type="solid")

    if data:
        headers = list(data[0].keys())
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for row_idx, row_data in enumerate(data, 2):
            for col_idx, key in enumerate(headers, 1):
                ws.cell(row=row_idx, column=col_idx, value=row_data.get(key, ""))

        # Auto-adjust column widths
        for col in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 40)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{title}.xlsx"'}
    )


def _draw_environmental_chart(data: list):
    from reportlab.graphics.shapes import Drawing, String
    from reportlab.graphics.charts.barcharts import VerticalBarChart
    from reportlab.lib import colors

    # Group by department
    dept_emissions = {}
    for r in data:
        dept = r.get("Department", "Unknown")
        try:
            emissions = float(r.get("Emission (kg CO2e)", 0))
        except ValueError:
            emissions = 0.0
        dept_emissions[dept] = dept_emissions.get(dept, 0.0) + emissions
    
    if not dept_emissions:
        return None

    # Limit to top 5 departments for rendering space
    sorted_depts = sorted(dept_emissions.items(), key=lambda x: x[1], reverse=True)[:5]
    labels = [d[0] for d in sorted_depts]
    values = [d[1] for d in sorted_depts]

    d = Drawing(400, 160)
    bc = VerticalBarChart()
    bc.x = 40
    bc.y = 20
    bc.height = 110
    bc.width = 320
    bc.data = [values]
    bc.categoryAxis.categoryNames = labels
    bc.categoryAxis.labels.fontSize = 7
    bc.categoryAxis.labels.dy = -10
    bc.valueAxis.valueMin = 0
    bc.valueAxis.valueMax = max(values) * 1.1 if values and max(values) > 0 else 100
    bc.valueAxis.valueStep = max(values) / 4 if values and max(values) > 0 else 25
    bc.valueAxis.labels.fontSize = 7
    bc.bars[0].fillColor = colors.HexColor("#10B981") # Brand Green
    
    d.add(bc)
    d.add(String(200, 145, "CO2 Emissions by Department (kg CO2e)", textAnchor='middle', fontSize=9, fontName='Helvetica-Bold'))
    return d


def _draw_social_chart(data: list):
    from reportlab.graphics.shapes import Drawing, String
    from reportlab.graphics.charts.piecharts import Pie
    from reportlab.lib import colors

    status_counts = {}
    for r in data:
        status = r.get("Status", "Pending")
        status_counts[status] = status_counts.get(status, 0) + 1
        
    if not status_counts:
        return None
        
    labels = list(status_counts.keys())
    values = list(status_counts.values())
    
    d = Drawing(400, 160)
    pc = Pie()
    pc.x = 120
    pc.y = 10
    pc.width = 120
    pc.height = 120
    pc.data = values
    pc.labels = [f"{l} ({v})" for l, v in zip(labels, values)]
    pc.sideLabels = 1
    
    color_palette = [colors.HexColor("#10B981"), colors.HexColor("#EF4444"), colors.HexColor("#F59E0B")]
    for i in range(len(values)):
        pc.slices[i].fillColor = color_palette[i % len(color_palette)]
        
    d.add(pc)
    d.add(String(200, 145, "CSR Participation Status Breakdown", textAnchor='middle', fontSize=9, fontName='Helvetica-Bold'))
    return d


def _draw_governance_chart(data: list):
    from reportlab.graphics.shapes import Drawing, String
    from reportlab.graphics.charts.piecharts import Pie
    from reportlab.lib import colors

    sev_counts = {}
    for r in data:
        sev = r.get("Severity", "Unknown")
        sev_counts[sev] = sev_counts.get(sev, 0) + 1
        
    if not sev_counts:
        return None
        
    labels = list(sev_counts.keys())
    values = list(sev_counts.values())
    
    d = Drawing(400, 160)
    pc = Pie()
    pc.x = 120
    pc.y = 10
    pc.width = 120
    pc.height = 120
    pc.data = values
    pc.labels = [f"{l} ({v})" for l, v in zip(labels, values)]
    pc.sideLabels = 1
    
    color_palette = [colors.HexColor("#EF4444"), colors.HexColor("#F59E0B"), colors.HexColor("#3B82F6"), colors.HexColor("#10B981")]
    for i in range(len(values)):
        pc.slices[i].fillColor = color_palette[i % len(color_palette)]
        
    d.add(pc)
    d.add(String(200, 145, "Compliance Issues by Severity", textAnchor='middle', fontSize=9, fontName='Helvetica-Bold'))
    return d


def _draw_esg_summary_chart(data: list):
    from reportlab.graphics.shapes import Drawing, String
    from reportlab.graphics.charts.barcharts import VerticalBarChart
    from reportlab.lib import colors

    dept_scores = {}
    for r in data:
        dept = r.get("Department", "Unknown")
        try:
            score = float(r.get("Total Score", 0))
        except ValueError:
            score = 0.0
        dept_scores[dept] = score
        
    if not dept_scores:
        return None
        
    sorted_depts = sorted(dept_scores.items(), key=lambda x: x[1], reverse=True)[:5]
    labels = [d[0] for d in sorted_depts]
    values = [d[1] for d in sorted_depts]
    
    d = Drawing(400, 160)
    bc = VerticalBarChart()
    bc.x = 40
    bc.y = 20
    bc.height = 110
    bc.width = 320
    bc.data = [values]
    bc.categoryAxis.categoryNames = labels
    bc.categoryAxis.labels.fontSize = 7
    bc.categoryAxis.labels.dy = -10
    bc.valueAxis.valueMin = 0
    bc.valueAxis.valueMax = 100
    bc.valueAxis.valueStep = 25
    bc.valueAxis.labels.fontSize = 7
    bc.bars[0].fillColor = colors.HexColor("#3B82F6") # Blue
    
    d.add(bc)
    d.add(String(200, 145, "Total ESG Scores by Department", textAnchor='middle', fontSize=9, fontName='Helvetica-Bold'))
    return d


def _generate_pdf(data: list, title: str):
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet

    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=landscape(letter))
    elements = []
    styles = getSampleStyleSheet()

    # Title
    elements.append(Paragraph(f"<b>{title}</b>", styles["Title"]))
    elements.append(Paragraph(f"Generated: {date.today()}", styles["Normal"]))
    elements.append(Spacer(1, 15))

    # Add Chart to PDF
    if data:
        chart_drawing = None
        if "Environmental" in title:
            chart_drawing = _draw_environmental_chart(data)
        elif "Social" in title:
            chart_drawing = _draw_social_chart(data)
        elif "Governance" in title:
            chart_drawing = _draw_governance_chart(data)
        elif "ESG Summary" in title:
            chart_drawing = _draw_esg_summary_chart(data)

        if chart_drawing:
            elements.append(chart_drawing)
            elements.append(Spacer(1, 15))

    if data:
        headers = list(data[0].keys())
        table_data = [headers]
        for row in data:
            table_data.append([str(row.get(h, "")) for h in headers])

        t = Table(table_data, repeatRows=1)
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1B5E20")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("FONTSIZE", (0, 0), (-1, 0), 9),
            ("FONTSIZE", (0, 1), (-1, -1), 8),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
        ]))
        elements.append(t)
    else:
        elements.append(Paragraph("No data found for the selected filters.", styles["Normal"]))

    doc.build(elements)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{title}.pdf"'}
    )
